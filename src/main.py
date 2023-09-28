import copy
import csv
import itertools
import os
import pathlib
import random
from typing import List, Callable

import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import time
import io
import tempfile

import matplotlib.pyplot as plt
from collections import Counter
import math
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt


class KilometricPoint:
    """ This symbolizes a single Kpoint and its corresponding coverages for each station """

    def __init__(self, point_val: float, coverages: list[str]):
        self.point_val = point_val
        self.coverages = coverages


class DataPoint:
    """ [This is redudant, for the sake of testing]
        Each plot point: KilometricPoint and its maximum coverage for set of active interest points. """

    def __init__(self, pk: KilometricPoint, max_coverage: float):
        self.pk = pk
        self.max_coverage = max_coverage


class LineCoverage:
    """ This symbolizes the coverage value along the railway line"""

    def __init__(self, points: list[str], pk_values: list[float]):
        self.max_coverages = [datapoint.max_coverage for datapoint in points]
        self.count = 0
        self.points_below_ref = self.compute_points_below_ref(pk_values)

    def compute_points_below_ref(self, pk_values: list[float]):
        points_below_ref = 0
        for i in range(len(self.max_coverages)):
            if low_signal_ref >= self.max_coverages[i] >= lim_min_coverage:
                points_below_ref += 1
            if self.max_coverages[i] < lim_min_coverage or math.isnan(self.max_coverages[i]):
                self.count += 1
        return points_below_ref

    def get_percent_below_min(self):
        return (self.count / len(self.max_coverages)) * 100

    def get_percent_above_min(self):
        return 100 - ((self.count / len(self.max_coverages)) * 100)

    def get_percent_below_ref(self):
        return (self.points_below_ref / len(self.max_coverages)) * 100

    @property
    def get_max_coverages(self):
        return self.max_coverages

    def get_max_coverages_public(self):
        return self.max_coverages


class Individual:
    """This symbolizes one individual - a random test case with random active/non-active stops."""

    def __init__(self, stopTypes: list[str]):
        self.stopTypes = stopTypes  # List of types of stops (Strings)
        # Using _ because nothing depends on that variable on the for
        self.individualStops = [random.randint(0, 1) for _ in stopTypes]
        self.id = 0
        self.verifyAnchors()
        self.value_cost_function = 0
        self.individual_weight = 0
        self.individual_weight_percent = 0
        self.percent_below_min = 0
        self.percent_below_ref = 0
        self.max_extension_low_ref = 0
        self.num_station = 0
        self.num_halt = 0
        self.num_anchor = 0
        self.num_signal = 0
        self.num_level_crossing = 0
        self.num_other = 0
        self.weight_percent_below_min = 55
        self.weight_percent_low_ref = 20
        self.weight_extension = 25
        self.weight_station = 2
        self.weight_halt = 3
        self.weight_anchor = 1
        self.weight_level_crossing = 5
        self.weight_signal = 7
        self.weight_other = 10
        self.individual_coverage = LineCoverage([], [])
        self.individual_pks = []

    def set_individualStops(self, individualStops: list[int]):
        self.individualStops = individualStops

    def get_Number_Active_Stops(self) -> int:
        return self.individualStops.count(1)  # Count the number of active stations

    def get_Number_Stops(self) -> int:
        return len(self.stopTypes)

    def verifyAnchors(self):
        # Guarantee that the Anchor is always active
        for index, stopType in enumerate(self.stopTypes):
            if stopType == 'Anchor':
                self.individualStops[index] = 1

    def compute_cost_function(self, stations_active_number: int, below_min: float, low_ref: float, extension: float,
                              list_stations: list[int],
                              stopTypes: list[str]):
        halt = 0
        station = 0
        anchor = 0
        level_crossing = 0
        signal = 0
        other = 0
        # weight_sum = self.weight_active_stations + self.weight_percent_below_min + self.weight_percent_low_ref + self.weight_extension
        total_stops = self.get_Number_Stops()

        # num_activeStations = self.get_Number_Active_Stops() # TODO: refactor this method: stations active number can be accessed from within the class

        for index, stop in enumerate(self.individualStops):
            if stop:  # Check if stop is active
                term = stopTypes[index]
                match term:
                    case 'Station':
                        station += 1
                    case 'Halt':
                        halt += 1
                    case 'Level Crossing':
                        level_crossing += 1
                    case 'Anchor':
                        anchor += 1
                    case 'Signal':
                        signal += 1
                    case 'Other':
                        other += 1
        if anchor == 0 and stopTypes.count('Anchor') != 0:
            print('Anchor station must be always active in every solution!')
            exit()
        self.num_station = station
        self.num_halt = halt
        self.num_signal = signal
        self.num_anchor = anchor
        self.num_level_crossing = level_crossing
        self.num_other = other
        best_value = ((2 * self.weight_station) / total_stops) + (
                (1 * self.weight_halt) / total_stops) + \
                     ((1 * self.weight_anchor) / total_stops) + (
                             (0 * self.weight_level_crossing) / total_stops) + \
                     ((0 * self.weight_signal) / total_stops) + (
                             (0 * self.weight_other) / total_stops) + \
                     (0 * self.weight_percent_below_min) + (
                             0 * self.weight_percent_low_ref) + (
                             0 * self.weight_extension)

        self.value_cost_function = ((self.num_station * self.weight_station) / total_stops) + (
                (self.num_halt * self.weight_halt) / total_stops) + \
                                   ((self.num_anchor * self.weight_anchor) / total_stops) + (
                                           (self.num_level_crossing * self.weight_level_crossing) / total_stops) + \
                                   ((self.num_signal * self.weight_signal) / total_stops) + (
                                           (self.num_other * self.weight_other) / total_stops) + \
                                   (below_min * self.weight_percent_below_min) + (
                                           low_ref * self.weight_percent_low_ref) + (
                                           extension * self.weight_extension)

        return self.value_cost_function

    @property
    def get_individualStops(self):
        return self.individualStops

    def __str__(self):
        return f"{self.individualStops}"


def read_coverages_file():
    stations = []
    pk_list = []
    pk_values = []
    file_path_coverage = '../data/Coverage_Cascais.csv'
    with open(file_path_coverage, 'r') as source_file:
        data_reader = csv.reader(source_file)

        for i, line in enumerate(data_reader):
            if i == 0:
                stations = [name for name in line]
                stations.pop(0)
            else:
                try:
                    pk_value = float(line[0])
                except ValueError:
                    print("Exception: Reached a pk that is not a float.")

                """ Append all coverage values. If the value is missing add a sufficiently low number instead. """
                pk_coverages = []
                for value in line:
                    try:
                        pk_coverages.append(float(value))
                    except ValueError:
                        if value == '':

                            pk_coverages.append(None)
                        else:
                            print(
                                "Exception: Reached a coverage that is not a float and is not an empty string either.")
                pk_coverages.pop(0)

                """ Create new KilometricPoint with its value and coverages for each interest point. """
                pk_list.append(KilometricPoint(pk_value, pk_coverages))
                pk_values.append(pk_value)
    return pk_values, stations, pk_list


def file_creation(filename1: str, filename2: str, filename3: str, filename4: str, generations_number: int):
    table_hearders_all_data = ["Individual", "Line Total Stops", "Cost Function Value", "Percentage Below Minimum",
                               "Weight",
                               "Percentage Low Reference", "Weight", "Maximum Extension Low Reference", "Weight",
                               "No. Active Sites", "No. Station", "Weight", "No. Halt", "Weight", "No. Anchors",
                               "Weight", "No. Signal", "Weight", "No. Level_Crossing", "Weight", "No. Other",
                               "Weight", "Variable Cost Function Value"]
    table_hearders_best_individuals = ["Generation", "Individual", "Cost Function Value", "Percentage Below Minimum",
                                       "Percentage Low Reference",
                                       "Maximum Extension Low Reference", "No. Active Sites", "No. Station", "No. Halt",
                                       "No. Anchors",
                                       "No. Signal", "No. Level_Crossing", "No. Other", "Graphic"]

    # Creating 2 new workbook
    workbook1 = openpyxl.Workbook()
    workbook2 = openpyxl.Workbook()
    workbook3 = openpyxl.Workbook()
    workbook4 = openpyxl.Workbook()

    # Creating the number of pages according to the number of generations
    for i in range(generations_number):
        sheet_name = "Generation n." + str(i + 1)
        worksheet1 = workbook1.create_sheet(sheet_name)
        worksheet3 = workbook3.create_sheet(sheet_name)

        for j, header in enumerate(table_hearders_all_data):
            worksheet1.cell(row=1, column=j + 1, value=header)
            worksheet3.cell(row=1, column=j + 1, value=header)

    sheet_name1 = "Cost_Function Evolution"
    worksheet2 = workbook2.create_sheet(sheet_name1)
    worksheet4 = workbook4.create_sheet(sheet_name1)
    sheet_name2 = "Cost_Function Generation"
    worksheet2 = workbook2.create_sheet(sheet_name2)
    worksheet4 = workbook4.create_sheet(sheet_name2)

    for i, header in enumerate(table_hearders_best_individuals):
        worksheet2.cell(row=1, column=i + 1, value=header)
        worksheet4.cell(row=1, column=i + 1, value=header)

    workbook1.remove(workbook1[workbook1.sheetnames[0]])
    workbook1.save(filename1)
    workbook2.remove(workbook2[workbook2.sheetnames[0]])
    workbook2.save(filename2)
    workbook3.remove(workbook3[workbook3.sheetnames[0]])
    workbook3.save(filename3)
    workbook4.remove(workbook4[workbook4.sheetnames[0]])
    workbook4.save(filename4)

    return workbook1, workbook2, workbook3, workbook4


def identify_station_pks(stations_names: list[str], active_stations: list[str]):
    interesting_pks = []
    file_path_pks = '../data/Cascais Elements_pks.csv'
    with open(file_path_pks, 'r') as source_file:
        data_reader2 = csv.reader(source_file)
        station_names_and_pks = [station for station in data_reader2]
        station_names_and_pks.pop(0)

        for indx, station_status in enumerate(active_stations):
            if (station_status):
                station_name_1 = stations_names[indx]
                for station_name_and_pk in station_names_and_pks:
                    if station_name_1 == station_name_and_pk[0]:
                        interesting_pks.append(station_name_and_pk[1])
    return interesting_pks


def processDataPoints(station_list: list[str], pks: list[KilometricPoint]):
    data_points = []
    for pk in pks:

        max_coverage = -1200  # If a pk has all values missing
        # np.max(pk.coverages[pk])
        for j, station in enumerate(station_list):  # Checks the highest coverage value
            if pk.coverages[j] is not None:
                if station and pk.coverages[j] > max_coverage:  # Station is active
                    max_coverage = pk.coverages[j]
        if max_coverage != -1200:
            data_points.append(DataPoint(pk, max_coverage))
        else:
            data_points.append(DataPoint(pk, float('nan')))
    return data_points


def processStationPoints(active_station_pks: list[str], datapoints: list[DataPoint]):
    for station_pk in active_station_pks:
        for indx, datapoint in enumerate(datapoints):
            if datapoint.pk.point_val == float(station_pk):
                adjacent_value = None
                offset = 1
                while (
                        adjacent_value is None):  # find the closest adjancent value, that actually has a value!
                    # Perhaps change this to find the max adjacent value instead...
                    direction = offset % 2
                    if direction == 1:
                        if indx + offset < len(datapoints):
                            try:
                                adjacent_value = datapoints[indx + offset].max_coverage  # check values on right
                            except:
                                print("index", indx)
                                print("offset", offset)
                                print("len(datapoints)", len(datapoints))
                        elif indx - offset >= 0:
                            adjacent_value = datapoints[indx - offset].max_coverage  # check values on right
                        else:
                            raise Exception
                    else:
                        if indx - offset >= 0:
                            adjacent_value = datapoints[indx - offset].max_coverage  # check values on left
                        elif indx + offset <= len(datapoints):
                            adjacent_value = datapoints[indx + offset].max_coverage  # check values on right
                        else:
                            raise Exception
                    offset += 1
                datapoint.max_coverage = adjacent_value
                break
            elif datapoint.pk.point_val > float(station_pk):
                break


def compute_max_extension(coverage_points: list[float], pk_reference: list[float]):
    point_first = None
    point_last = None
    max_extension = 0
    for index, coverage_point in enumerate(coverage_points):
        if lim_min_coverage <= coverage_point < low_signal_ref:  # inside limits
            if point_first is None:
                point_first = pk_reference[index]
            if point_first is not None and index + 1 >= len(coverage_points):  # in case its the last point
                point_last = pk_reference[index]
                max_temp = compute_extension_distance(point_first, point_last)
                max_extension = max_temp if max_temp > max_extension else max_extension
        if (lim_min_coverage > coverage_point or coverage_point >= low_signal_ref) and (
                point_first is not None):  # outside limits
            point_last = pk_reference[index - 1]
            max_temp = compute_extension_distance(point_first, point_last)
            max_extension = max_temp if max_temp > max_extension else max_extension
            point_first = None
            point_last = None

    return max_extension


def compute_extension_distance(point_first: float, point_last: float):
    max_distance_temp = point_last - point_first
    max_distance = 0
    if max_distance_temp > max_distance:
        max_distance = max_distance_temp
    return max_distance


def read_stopTypes():
    types = []
    file_path_stopTypes = '../data/Cascais Elements_pks.csv'
    with open(file_path_stopTypes, 'r') as types_file:
        data_reader = csv.reader(types_file)
        next(data_reader)
        for row in data_reader:
            types.append(row[2])
    return types


def create_population(size_population: int, stop_types: list[str], existing_population: list[Individual]):
    # Generates the number os individuals according to the size of the population
    individualsStops = [str(individual.get_individualStops) for individual in existing_population]
    active_station_dict = dict.fromkeys(individualsStops, len(individualsStops))
    population = []
    i = 0
    while i < size_population:
        individual = Individual(stop_types)
        individualStops = individual.get_individualStops
        if str(individualStops) not in active_station_dict:
            active_station_dict[str(individualStops)] = i
            population.append(individual)
            i += 1
        individual.id = i

    return population


def compute_population_data(population: list[Individual]):
    for individual in population:
        num_active_stops = individual.get_Number_Active_Stops()
        individualStops = individual.get_individualStops
        activeStops_pks = identify_station_pks(stations, individualStops)
        datapoints = processDataPoints(individualStops, pk_list)
        processStationPoints(activeStops_pks, datapoints)
        line_coverage = LineCoverage(datapoints, pk_values)

        percent_below_min = line_coverage.get_percent_below_min()
        percent_below_ref = line_coverage.get_percent_below_ref()
        # percent_above_min = line_coverage.get_percent_above_min()
        max_coverages = line_coverage.get_max_coverages

        # Compute the consecutive max entension between the pk values below the minimum limit
        max_extension = compute_max_extension(max_coverages, pk_values)

        # Compute the cost function value for the individual
        individual_cost_function_value = individual.compute_cost_function(num_active_stops, percent_below_min,
                                                                          percent_below_ref, max_extension, stations,
                                                                          stop_Types)
        individual.percent_below_min = percent_below_min
        individual.percent_below_ref = percent_below_ref
        individual.max_extension_low_ref = max_extension
        individual.individual_coverage = line_coverage
        individual.individual_pks = pk_values
    return population


def rank_individuals(population: list[Individual]):
    return population.sort(key=lambda individual: individual.value_cost_function, reverse=False)


def roulette_selection(selection_population: list[Individual], eliteSize: int, crossover_probability: float,
                       population_size: int):
    population_total_cost_val_sum = 0
    cumulative_weights = [0]
    parents = []
    j = 0
    selection_number = round(crossover_probability * population_size)
    for i in range(eliteSize):
        selection_population.remove(selection_population[i])
    for individual in selection_population:
        population_total_cost_val_sum += individual.value_cost_function
        weights = 1 / individual.value_cost_function
        cumulative_weights.append(cumulative_weights[j] + weights)
        #individual.individual_weight_percent = (cumulative_weights[j + 1] / population_total_cost_val_sum) * 100
        j += 1
    cumulative_weights.pop(0)

    max_value = max(cumulative_weights)
    count = 0
    aux = [individual for individual in selection_population]
    for _ in range(selection_number):
        random_number = random.uniform(0, max_value)
        for score in cumulative_weights:
            if random_number <= score:
                parents.append(aux[count])
                selection_population.remove(selection_population[count])
                cumulative_weights.remove(score)
                break
        count += 1

    return parents


def tournament_selection(selection_population: list[Individual], eliteSize: int, crossover_probability: float,
                         population_size: int):
    parents = []
    elite = []

    # Store elite individuals separately
    elite = selection_population[:eliteSize]

    # Remove elite individuals from the selection population
    selection_population = selection_population[eliteSize:]

    selection_number = round(crossover_probability * len(selection_population))
    for _ in range(selection_number):
        if len(selection_population) > 1 :
            # Randomly select individuals for the tournament
            tournament_individuals = random.sample(selection_population, 2)

            # Find the individual with the lowest cost function value in the tournament
            winner = min(tournament_individuals, key=lambda x: x.value_cost_function)

            # Add the winner to the parents list
            parents.append(winner)

            # Remove the winner from the selection population
            selection_population.remove(winner)
        else:
            winner = selection_population[0]
            parents.append(winner)
    # Append the elite individuals to the parents list
    parents.extend(elite)

    return parents


def crossover(parents: list[Individual], crossover_probability: float):
    crossover_population = []
    random.shuffle(parents)
    parents_check = []
    parents_check.extend(parents)
    # Count the number of pairs for crossover
    number_pairs = math.floor((len(parents) / 2))

    if crossover_probability > 0:
        # Get the size of a chromossome
        chromossome_size = len(parents[0].individualStops)

        for pair in range(number_pairs):
            length = len(parents)
            first_parent_index = random.randrange(0, length)
            second_parent_index = random.randrange(0, length)
            while first_parent_index == second_parent_index:
                second_parent_index = random.randrange(0, length)

            stop = round(chromossome_size / 2)

            # Create child chromosomes by combining sections of the parents' chromosomes
            first_parent = parents[first_parent_index]
            second_parent = parents[second_parent_index]
            first_child = Individual(list(first_parent.individualStops))
            first_child.set_individualStops(first_parent.individualStops[0:stop] + second_parent.individualStops[stop:])
            second_child = Individual(list(second_parent.individualStops))
            second_child.set_individualStops(
                second_parent.individualStops[0:stop] + first_parent.individualStops[stop:])

            # Remove the parents from the pool
            parents.remove(first_parent)
            parents.remove(second_parent)

            crossover_population.append(first_child)
            crossover_population.append(second_child)
            crossover_population = compute_population_data(crossover_population)

    return crossover_population


def mutation(mutation_population: list[Individual], mutation_probability: float):
    # This method will apply a bit swap to one of the individuals on the crossover population
    population = []
    population.extend(mutation_population)
    rand_numb = 0
    for individual in mutation_population:
        while rand_numb == 0:
            rand_numb = round(random.uniform(0, 1), 2)
        if rand_numb <= mutation_probability:
            index_1 = random.randrange(len(individual.individualStops))
            while individual.stopTypes[index_1] == 'Anchor':
                index_1 = random.randrange(len(individual.individualStops))
            if individual.individualStops[index_1] == 1:
                individual.individualStops[index_1] = 0
            if individual.individualStops[index_1] == 0:
                individual.individualStops[index_1] = 1

    return mutation_population


def generations_creation(population_size: int, stop_Types: list[str], number_generations: int, eliteSize: int,
                         crossover_probability: float, mutation_probability: float, roulette_data_file: Workbook,
                         tournament_data_file: Workbook, roulette_filename: str, tournament_filename: str):
    best_individuals = []
    elite_individuals = []
    initial_population = create_population(population_size, stop_Types, [])
    computed_population = compute_population_data(initial_population)
    rank_individuals(computed_population)
    all_populations_data = []
    index = 0

    def run_selection(selection_method: Callable, best_individuals: list, elite_individuals: list,
                      population: list[Individual], computed_population: list[Individual], all_populations_data: list):

        for generation in range(number_generations):
            print("Generation no:", generation)
            number_elite = len(elite_individuals)
            if number_elite == 0:
                elite = get_elite(computed_population, eliteSize)
                parents = selection_method(computed_population, eliteSize, crossover_probability, population_size)

            new_population_size = population_size - number_elite
            if number_elite > 0:
                elite = get_elite(final_population, eliteSize)
                parents = selection_method(final_population, eliteSize, crossover_probability, population_size)
            if number_elite < 0:
                print("Elite can never be lower than 0!")
                exit()
            crossover_population = crossover(parents, crossover_probability)
            after_crossover_population = crossover_population + elite
            filled_population = fill_population(after_crossover_population, population_size, stop_Types)
            final_population = mutation(filled_population, mutation_probability)
            final_population = compute_population_data(final_population)
            rank_individuals(final_population)
            elite_individuals = get_elite(final_population, eliteSize)
            best_individuals.append(final_population[0])
            if selection_method.__name__ == "roulette_selection":
                all_data(final_population, generation, roulette_data_file, roulette_filename)
            if selection_method.__name__ == "tournament_selection":
                all_data(final_population, generation, tournament_data_file, tournament_filename)
            generation_data = []
            for index in range(len(final_population)):
                generation_data.append(final_population[index])
            all_populations_data.append(generation_data)
            parents = []
            after_crossover_population = []
        returning_best_individuals = []
        returning_all_populations_data = []
        returning_best_individuals[:] = best_individuals
        best_individuals.clear()
        returning_all_populations_data[:] = all_populations_data
        all_populations_data.clear()

        return returning_best_individuals, returning_all_populations_data

    # Run with the tournament selection method
    tournament_best_individuals, tournament_all_populations_data = run_selection(tournament_selection, best_individuals,
                                                                                 elite_individuals, initial_population,
                                                                                 computed_population,
                                                                                 all_populations_data)

    computed_population = compute_population_data(initial_population)

    # Run with the roulette selection method
    roulette_best_individuals, roulette_all_populations_data = run_selection(roulette_selection, best_individuals,
                                                                             elite_individuals, initial_population,
                                                                             computed_population, all_populations_data)

    return roulette_best_individuals, roulette_all_populations_data, tournament_best_individuals, tournament_all_populations_data


def fill_population(new_population: list[Individual], population_size: int, stop_Types: list[str]):
    this_population_size = population_size - len(new_population)
    filling_population = create_population(this_population_size, stop_Types, new_population)
    filling_population = compute_population_data(filling_population)
    new_population = new_population + filling_population

    return new_population


def get_elite(new_population: list[Individual], eliteSize: int):
    # Getting the elite individuals to go into the next gen
    elite_individuals = []
    number_elite = 0
    for individual in new_population:
        if number_elite < eliteSize:
            # Usage of deepcopy to create a new (equal) object
            elite_individuals.append(copy.deepcopy(individual))
            number_elite += 1
        else:
            break
    return elite_individuals


def all_data(new_population: list[Individual], generation: int, all_data_file: Workbook, filename: str):
    sheet_name1 = "Generation n." + str(generation + 1)
    worksheet = all_data_file[sheet_name1]

    for k, individual in enumerate(new_population):
        row = k + 2  # Offset by 1 to account for headers
        worksheet.cell(row=row, column=1, value=str(individual.individualStops))
        worksheet.cell(row=row, column=2, value=float(individual.get_Number_Stops()))
        worksheet.cell(row=row, column=3, value=float(individual.value_cost_function))
        worksheet.cell(row=row, column=4, value=float(individual.percent_below_min))
        worksheet.cell(row=row, column=5, value=float(individual.weight_percent_below_min))
        worksheet.cell(row=row, column=6, value=float(individual.percent_below_ref))
        worksheet.cell(row=row, column=7, value=float(individual.weight_percent_low_ref))
        worksheet.cell(row=row, column=8, value=float(individual.max_extension_low_ref))
        worksheet.cell(row=row, column=9, value=float(individual.weight_extension))
        worksheet.cell(row=row, column=10, value=individual.get_Number_Active_Stops())
        worksheet.cell(row=row, column=11, value=individual.num_station)
        worksheet.cell(row=row, column=12, value=individual.weight_station)
        worksheet.cell(row=row, column=13, value=individual.num_halt)
        worksheet.cell(row=row, column=14, value=individual.weight_halt)
        worksheet.cell(row=row, column=15, value=individual.num_anchor)
        worksheet.cell(row=row, column=16, value=individual.weight_anchor)
        worksheet.cell(row=row, column=17, value=individual.num_signal)
        worksheet.cell(row=row, column=18, value=individual.weight_signal)
        worksheet.cell(row=row, column=19, value=individual.num_level_crossing)
        worksheet.cell(row=row, column=20, value=individual.weight_level_crossing)
        worksheet.cell(row=row, column=21, value=individual.num_other)
        worksheet.cell(row=row, column=22, value=individual.weight_other)
        cost_function_formula = f'(((K{row} * L{row})/ B{row}) + ((M{row} * N{row}) / B{row}) + ((O{row} * P{row}) / B{row}) + ((Q{row} * R{row}) / B{row})  + ((S{row} * T{row}) / B{row}) + ((U{row} * V{row}) / B{row}) + (D{row} * E{row}) + (F{row} * G{row}) + (H{row} * I{row}))'

        formula = "=" + cost_function_formula
        worksheet.cell(row=row, column=24).value = formula
        # worksheet.cell(row=row, column=22, value=cost_function_formula.format(row=row))
    # workbook3.save(filename)
    all_data_file.save(filename)


def results_data(roulette_best_individuals: list[Individual], tournament_best_individuals: list[Individual],
                 roulette_best_individuals_file: Workbook, tournament_best_individuals_file: Workbook,
                 generation: list[int], roulette_best_individual_cost_value: list[int],
                 tournament_best_individual_cost_value: list[int], population_size: int, number_generations: int,
                 crossover_prob: float, mutation_prob: float, roulette_results_filename: str,
                 tournament_results_filename: str):
    sheet_name1 = "Cost_Function Evolution"
    worksheet1 = roulette_best_individuals_file[sheet_name1]
    fig1, ax1 = plt.subplots()
    generation_evolution_plot(generation, roulette_best_individual_cost_value)

    # Convert plot to image
    image_stream1 = io.BytesIO()
    plt.savefig(image_stream1, format='png')
    plt.close(fig1)

    # Insert the image into the Excel worksheet
    image_stream1.seek(0)
    img1 = Image(image_stream1)
    img1.width = 300  # Adjust the width of the image if needed
    img1.height = 220  # Adjust the height of the image if needed
    worksheet1.add_image(img1, f'C{4}')
    worksheet1.cell(row=5, column=10, value=f"No. Individuals: {population_size}")
    worksheet1.cell(row=6, column=10, value=f"No. Generations: {number_generations}")
    worksheet1.cell(row=7, column=10, value=f"Crossover Prob.: {crossover_prob}")
    worksheet1.cell(row=8, column=10, value=f"Mutation Prob.: {mutation_prob}")

    worksheet3 = tournament_best_individuals_file[sheet_name1]
    fig2, ax2 = plt.subplots()
    generation_evolution_plot(generation, tournament_best_individual_cost_value)
    # Convert plot to image
    image_stream2 = io.BytesIO()
    plt.savefig(image_stream2, format='png')
    plt.close(fig2)

    # Insert the image into the Excel worksheet
    image_stream2.seek(0)
    img2 = Image(image_stream2)
    img2.width = 300  # Adjust the width of the image if needed
    img2.height = 220  # Adjust the height of the image if needed
    worksheet3.add_image(img2, f'C{4}')
    worksheet3.cell(row=5, column=10, value=f"No. Individuals: {population_size}")
    worksheet3.cell(row=6, column=10, value=f"No. Generations: {number_generations}")
    worksheet3.cell(row=7, column=10, value=f"Crossover Prob.: {crossover_prob}")
    worksheet3.cell(row=8, column=10, value=f"Mutation Prob.: {mutation_prob}")

    sheet_name2 = "Cost_Function Generation"
    worksheet2 = roulette_best_individuals_file[sheet_name2]

    # Loop for roulette method results
    for k, individual in enumerate(roulette_best_individuals):
        row = k + 2
        worksheet2.cell(row=row, column=1, value=k + 1)
        worksheet2.cell(row=row, column=2, value=str(individual.individualStops))
        worksheet2.cell(row=row, column=3, value=float(individual.value_cost_function))
        worksheet2.cell(row=row, column=4, value=float(individual.percent_below_min))
        worksheet2.cell(row=row, column=5, value=float(individual.percent_below_ref))
        worksheet2.cell(row=row, column=6, value=float(individual.max_extension_low_ref))
        worksheet2.cell(row=row, column=7, value=individual.get_Number_Active_Stops())
        worksheet2.cell(row=row, column=8, value=individual.num_station)
        worksheet2.cell(row=row, column=9, value=individual.num_halt)
        worksheet2.cell(row=row, column=10, value=individual.num_anchor)
        worksheet2.cell(row=row, column=11, value=individual.num_signal)
        worksheet2.cell(row=row, column=12, value=individual.num_level_crossing)
        worksheet2.cell(row=row, column=13, value=individual.num_other)

        # Generate the plot
        fig, ax = plt.subplots()
        cov_data_plot(generation=k, best_individuals=roulette_best_individuals)

        # Convert plot to image
        image_stream1 = io.BytesIO()
        plt.savefig(image_stream1, format='png')
        plt.close(fig)

        # Insert the image into the Excel worksheet (Column 14 - Column N)
        image_stream1.seek(0)
        img1 = Image(image_stream1)
        worksheet2.column_dimensions['N'].width = 40
        worksheet2.row_dimensions[row].height = 170
        img1.width = 300  # Adjust the width of the image if needed
        img1.height = 220  # Adjust the height of the image if needed
        worksheet2.add_image(img1, f'N{row}')

        # worksheet.cell(row=row, column=14, value=cov_data_plot(k, best_individuals))
    worksheet4 = tournament_best_individuals_file[sheet_name2]
    # Loop for tournament method results
    for i, individual in enumerate(tournament_best_individuals):
        row = i + 2
        worksheet4.cell(row=row, column=1, value=i + 1)
        worksheet4.cell(row=row, column=2, value=str(individual.individualStops))
        worksheet4.cell(row=row, column=3, value=float(individual.value_cost_function))
        worksheet4.cell(row=row, column=4, value=float(individual.percent_below_min))
        worksheet4.cell(row=row, column=5, value=float(individual.percent_below_ref))
        worksheet4.cell(row=row, column=6, value=float(individual.max_extension_low_ref))
        worksheet4.cell(row=row, column=7, value=individual.get_Number_Active_Stops())
        worksheet4.cell(row=row, column=8, value=individual.num_station)
        worksheet4.cell(row=row, column=9, value=individual.num_halt)
        worksheet4.cell(row=row, column=10, value=individual.num_anchor)
        worksheet4.cell(row=row, column=11, value=individual.num_signal)
        worksheet4.cell(row=row, column=12, value=individual.num_level_crossing)
        worksheet4.cell(row=row, column=13, value=individual.num_other)

        # Generate the plot
        fig, ax = plt.subplots()
        cov_data_plot(generation=i, best_individuals=tournament_best_individuals)

        # Convert plot to image
        image_stream2 = io.BytesIO()
        plt.savefig(image_stream2, format='png')
        plt.close(fig)

        # Insert the image into the Excel worksheet (Column 14 - Column N)
        image_stream2.seek(0)
        img2 = Image(image_stream2)
        worksheet4.column_dimensions['N'].width = 40
        worksheet4.row_dimensions[row].height = 170
        img2.width = 300  # Adjust the width of the image if needed
        img2.height = 220  # Adjust the height of the image if needed
        worksheet4.add_image(img2, f'N{row}')

    roulette_best_individuals_file.save(roulette_results_filename)
    tournament_best_individuals_file.save(tournament_results_filename)


def cov_data_plot(generation: int, best_individuals: list[Individual]):
    # for generation in range(number_generations):
    pk_axis = best_individuals[generation].individual_pks
    coverage_axis_aux = best_individuals[generation].individual_coverage
    coverage_axis = coverage_axis_aux.get_max_coverages
    plt.figure()
    # plt.figure(1)
    plt.plot(pk_axis, coverage_axis, color='black')
    plt.xlim(0, max(pk_values))
    plt.axhline(y=lim_min_coverage, color='r')
    plt.axhline(y=low_signal_ref, color='y')
    plt.xlabel("Distance (pk)")
    plt.ylabel("Coverage (dBm)")
    # plt.title("Coverage Map Bad Individual")
    title = f"Coverage Map of the best individual of Generation number {generation + 1}"
    # plt.show()
    plt.title(title)


def generation_evolution_plot(generation: list[int], best_individual_cost_value: list[int]):
    xtick = 1
    if 100 > max(generation) >= 30:
        xtick = round(max(generation) / (max(generation) / 2))
    if 30 < max(generation) >= 100:
        xtick = round(max(generation) / (max(generation) / 10))

    plt.figure(1)
    plt.plot(generation, best_individual_cost_value)
    plt.xlabel("Generation")
    plt.ylabel("Best individual cost function value")
    plt.title("Cost Function Value evolution of the best individual")
    plt.xticks(range(0, max(generation) + 1, xtick))


if __name__ == "__main__":
    start_time = time.time()
    filename1 = "Roulette_data_storage.xlsx"
    filename2 = "Roulette_Result_Test.xlsx"
    filename3 = "Tournament_data_storage.xlsx"
    filename4 = "Tournament_Result_Test.xlsx"
    lim_min_coverage = -95
    low_signal_ref = -85
    # This will be the initial size, the population after selection crossover and
    # mutation will have a random size based on the selection occurance
    number_generations = 10
    population_size = 10
    crossover_probability = 0.5
    mutation_probability = 0.01
    roulette_best_individuals = []
    all_cost_function_data = []

    # Create the xlsx file to store the data from all the generations
    workbook1, workbook2, workbook3, workbook4 = file_creation(filename1, filename2, filename3, filename4,
                                                               number_generations)

    # Size of elite population should never be higher than 1% of the population size
    eliteSize = round(0.01 * population_size)
    if eliteSize == 0:
        eliteSize = 1

    # Read external files
    pk_values, stations, pk_list = read_coverages_file()
    stop_Types = read_stopTypes()

    roulette_best_individuals, roulette_all_population, tournament_best_individuals, tournament_all_population = \
        generations_creation(population_size, stop_Types, number_generations, eliteSize, crossover_probability,
                             mutation_probability, workbook1, workbook3, filename1, filename3)

    roulette_best_individual_cost_value = []
    tournament_best_individual_cost_value = []
    generation = []
    gen = 1
    for individual in roulette_best_individuals:
        roulette_best_individual_cost_value.append(individual.value_cost_function)
        generation.append(gen)
        gen += 1

    for individual in tournament_best_individuals:
        tournament_best_individual_cost_value.append(individual.value_cost_function)

    results_data(roulette_best_individuals, tournament_best_individuals, workbook2, workbook4, generation,
                 roulette_best_individual_cost_value, tournament_best_individual_cost_value, population_size,
                 number_generations, crossover_probability, mutation_probability, filename2, filename4)

    end_time = time.time()
    elapsed_time = end_time - start_time
    print("Elapsed time: ", elapsed_time)
